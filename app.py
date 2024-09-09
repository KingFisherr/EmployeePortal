from flask import Flask, request, render_template, redirect
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
EXCEL_FILE = 'employee_data.xlsx'

# Initialize Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Employee ID", "Date", "Check-In Time", "Check-Out Time"])  # Adding headers
    wb.save(EXCEL_FILE)

# Route for Check-In Page
@app.route('/')
def index():
    return render_template('index.html')

# Combined route to handle Check-In and Check-Out
@app.route('/attendance', methods=['POST'])
def attendance():
    employee_id = request.form['employee_id']
    action = request.form['action']
    current_time = datetime.now().strftime("%I:%M %p")  # Format time as "HH:MM AM/PM"
    current_date = datetime.now().strftime("%Y-%m-%d")  # Format date as "YYYY-MM-DD"

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        if action == 'checkin':
            # Check-In: Add a new row with check-in time
            ws.append([employee_id, current_date, current_time, ""])  # Leave check-out time empty for now

        elif action == 'checkout':
            # Check-Out: Find the matching row and update the check-out time
            for row in ws.iter_rows(min_row=2):  # Assuming first row is a header
                if row[0].value == employee_id and row[1].value == current_date and not row[3].value:
                    row[3].value = current_time  # Update check-out time
                    break

        wb.save(EXCEL_FILE)

    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return "An error occurred while saving to Excel.", 500

    return redirect('/')

# Route for Status Page
@app.route('/status', methods=['GET', 'POST'])
def status():
    records = []
    if request.method == 'POST':
        employee_id = request.form['employee_id']
        
        # Read from Excel to find records for the given employee ID
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] == employee_id:
                records.append({
                    "date": row[1],
                    "checkin": row[2],
                    "checkout": row[3] if row[3] else "Not Checked Out"
                })

    return render_template('status.html', records=records)

if __name__ == '__main__':
    app.run(debug=True)
